"""Read Excel skill - reads data from Excel files."""

import openpyxl
from pathlib import Path

from .responses import ReadExcelResponse


async def main(file_path: str, sheet_name: str | None = None) -> ReadExcelResponse:
    """
    Read data from Excel file.

    Args:
        file_path: Path to Excel file (e.g., "/inputs/data.xlsx")
        sheet_name: Optional sheet name (defaults to active sheet)

    Returns:
        ReadExcelResponse with data, sheet_name, num_rows, num_cols
    """
    try:
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            # List available files in the directory to help debug
            directory = file_path_obj.parent
            available_files = []
            if directory.exists() and directory.is_dir():
                available_files = [f.name for f in directory.iterdir() if f.is_file()]

            error_msg = f"Excel file not found: {file_path}"
            if available_files:
                error_msg += f". Available files in {directory}: {', '.join(available_files)}"
            else:
                error_msg += f". Directory {directory} does not exist or is empty."

            return ReadExcelResponse(success=False, error=error_msg)

        workbook = openpyxl.load_workbook(file_path, data_only=True)

        # Always include available sheets in response
        available_sheets = workbook.sheetnames

        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                return ReadExcelResponse(
                    success=False,
                    error=f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(available_sheets)}",
                    available_sheets=available_sheets,
                )
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        if sheet is None:
            return ReadExcelResponse(
                success=False,
                error=f"No active worksheet found. Available sheets: {', '.join(available_sheets)}",
                available_sheets=available_sheets,
            )

        data = []
        for row in sheet.iter_rows(values_only=True):
            row_data = [
                ""
                if cell is None
                else str(cell)
                if not isinstance(cell, (str, int, float, bool))
                else cell
                for cell in row
            ]
            data.append(row_data)

        return ReadExcelResponse(
            success=True,
            sheet_name=sheet.title,
            available_sheets=available_sheets,
            num_rows=len(data),
            num_cols=len(data[0]) if data else 0,
            data=data,
        )

    except Exception as e:
        return ReadExcelResponse(success=False, error=f"Error reading Excel: {str(e)}")
